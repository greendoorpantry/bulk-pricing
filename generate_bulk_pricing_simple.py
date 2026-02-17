#!/usr/bin/env python3
"""
Bulk Pricing Data Generator - Lightweight Version
Works without numpy/pandas, uses only openpyxl
Auto-commits and pushes to GitHub after generation.
"""

import json
from datetime import datetime
import os
import subprocess
import sys
import io

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed")
    print("Run: pip install openpyxl")
    exit(1)

# ============================================================================
# CONFIGURATION
# ============================================================================

INPUT_FILE = "data/products.xlsx"  # Changed to .xlsx
OUTPUT_FILE = "data/bulk_pricing_data.json"
GOOGLE_DRIVE_PATH = ""

# CES Touch export settings
CES_EXPORT_FILE = r"C:\Touch\IMP-EXP\sku_0002.xls"

SHIPPING_PERCENTAGE = 0.07
YOUR_PROFIT_SHARE = 0.60
CUSTOMER_PROFIT_SHARE = 0.40
MINIMUM_DISCOUNT = 10.0
MAXIMUM_DISCOUNT = 25.0

# Git settings
GIT_ENABLED = True
REPO_DIR = r"C:\BulkPricing"

# ============================================================================

def round_to_nearest_10_cents(price):
    """Round price UP to nearest 10 cents (EUR 0.10)"""
    import math
    return math.ceil(price * 10) / 10


def import_from_ces():
    """Export from CES Touch via GUI automation, then convert .xls to .xlsx"""
    
    print(f"\n[IMPORT] Importing from CES Touch...")
    
    # Step 1: Run the GUI automation to export from CES Touch
    print("   [INFO] Running CES Touch GUI export...")
    try:
        # Import and run the export directly
        sys.path.insert(0, REPO_DIR)
        from ces_export import export_products
        
        success = export_products()
        if not success:
            print("   [ERROR] CES Touch export failed")
            return False
    except ImportError:
        print("   [ERROR] ces_export.py not found in C:\\BulkPricing")
        return False
    except Exception as e:
        print(f"   [ERROR] CES export error: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    # Step 2: Verify the export file exists
    if not os.path.exists(CES_EXPORT_FILE):
        print(f"   [ERROR] CES export not found: {CES_EXPORT_FILE}")
        return False
    
    # Show file age
    mod_time = datetime.fromtimestamp(os.path.getmtime(CES_EXPORT_FILE))
    age = datetime.now() - mod_time
    print(f"   [INFO] Export file last modified: {mod_time.strftime('%Y-%m-%d %H:%M')}")
    if age.total_seconds() > 300:  # More than 5 minutes old
        print(f"   [WARN] Export file seems stale - may not have been updated")
    
    # Step 3: Convert .xls to .xlsx
    os.makedirs("data", exist_ok=True)
    
    try:
        import win32com.client
        
        print("   [INFO] Converting .xls to .xlsx via Excel...")
        
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        xls_path = os.path.abspath(CES_EXPORT_FILE)
        xlsx_path = os.path.abspath(INPUT_FILE)
        
        wb = excel.Workbooks.Open(xls_path)
        wb.SaveAs(xlsx_path, FileFormat=51)
        wb.Close(False)
        excel.Quit()
        
        print(f"   [OK] Converted to: {INPUT_FILE}")
        return True
        
    except ImportError:
        print("   [WARN] win32com not available, trying xlrd fallback...")
        
        try:
            import xlrd
            from openpyxl import Workbook
            
            xls_wb = xlrd.open_workbook(CES_EXPORT_FILE)
            xls_ws = xls_wb.sheet_by_index(0)
            
            xlsx_wb = Workbook()
            xlsx_ws = xlsx_wb.active
            
            for row_idx in range(xls_ws.nrows):
                for col_idx in range(xls_ws.ncols):
                    xlsx_ws.cell(row=row_idx + 1, column=col_idx + 1, 
                               value=xls_ws.cell_value(row_idx, col_idx))
            
            xlsx_wb.save(INPUT_FILE)
            print(f"   [OK] Converted via xlrd to: {INPUT_FILE}")
            return True
            
        except ImportError:
            print("   [ERROR] Neither win32com nor xlrd available.")
            print("   Install one of:")
            print("     pip install pywin32")
            print("     pip install xlrd")
            return False
    
    except Exception as e:
        print(f"   [ERROR] Excel conversion failed: {e}")
        return False


def calculate_bulk_pricing(cost, case_qty, retail_price):
    """Calculate bulk pricing"""
    true_case_cost = cost * case_qty * (1 + SHIPPING_PERCENTAGE)
    individual_value = retail_price * case_qty
    profit_pool = individual_value - true_case_cost
    
    customer_share = profit_pool * CUSTOMER_PROFIT_SHARE
    discount_pct = (customer_share / individual_value) * 100
    discount_pct = max(MINIMUM_DISCOUNT, min(MAXIMUM_DISCOUNT, discount_pct))
    
    bulk_price = individual_value * (1 - discount_pct / 100)
    
    # Round bulk price UP to nearest 10 cents
    bulk_price_rounded = round_to_nearest_10_cents(bulk_price)
    customer_saves = individual_value - bulk_price_rounded
    
    # Recalculate discount percentage based on rounded price
    discount_pct_actual = (customer_saves / individual_value) * 100
    
    # Tax mode - also rounded
    tax_bulk_price = individual_value * 0.90
    tax_bulk_price_rounded = round_to_nearest_10_cents(tax_bulk_price)
    tax_customer_saves = individual_value - tax_bulk_price_rounded
    
    return {
        'bulkPrice': bulk_price_rounded,
        'customerSaves': round(customer_saves, 2),
        'discountPct': round(discount_pct_actual, 1),
        'taxBulkPrice': tax_bulk_price_rounded,
        'taxCustomerSaves': round(tax_customer_saves, 2),
        'taxDiscountPct': round((tax_customer_saves / individual_value) * 100, 1)
    }


def load_products(file_path):
    """Load products from XLSX"""
    print(f"\n[FILE] Reading: {file_path}")
    
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    col_map = {h.lower(): i for i, h in enumerate(headers) if h}
    
    print(f"   [OK] Found {len(headers)} columns")
    
    # Read products
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        
        try:
            supp = str(row[col_map.get('supp', -1)] or '').strip()
            
            # Include all suppliers, not just UDEA
            cost = float(row[col_map.get('cost', -1)] or 0)
            caseqty = int(row[col_map.get('caseqty', -1)] or 0)
            nprice1 = float(row[col_map.get('nprice1', -1)] or 0)
            
            if cost > 0 and caseqty >= 1 and caseqty < 1000 and nprice1 > 0:
                # Normalize barcode - strip whitespace and convert to string
                raw_plu = row[col_map.get('plu', -1)]
                if raw_plu is None:
                    plu = ''
                elif isinstance(raw_plu, (int, float)):
                    plu = str(int(raw_plu))  # Convert number to string, remove decimals
                else:
                    plu = str(raw_plu).strip()  # Strip any whitespace
                
                products.append({
                    'plu': plu,
                    'desc': str(row[col_map.get('desc', -1)] or ''),
                    'supp': supp,
                    'suppcode': str(row[col_map.get('suppcode', -1)] or ''),
                    'cost': cost,
                    'caseqty': caseqty,
                    'nprice1': nprice1
                })
        except:
            continue
    
    wb.close()
    print(f"   [OK] Loaded {len(products)} valid products")
    return products


def git_push(repo_dir, file_path):
    """Commit and push the updated JSON to GitHub"""
    print(f"\n[PUSH] Pushing to GitHub...")
    
    try:
        # Run all git commands from the repo directory
        def git(*args):
            result = subprocess.run(
                ['git'] + list(args),
                cwd=repo_dir,
                capture_output=True,
                text=True,
                timeout=60
            )
            if result.returncode != 0:
                print(f"   [WARN] git {args[0]}: {result.stderr.strip()}")
            return result
        
        # Pull latest first to avoid conflicts
        git('pull', '--rebase', 'origin', 'main')
        
        # Stage the JSON file
        git('add', file_path)
        
        # Check if there are changes to commit
        status = git('diff', '--cached', '--quiet')
        if status.returncode == 0:
            print("   [INFO] No changes to commit (data unchanged)")
            return True
        
        # Commit with timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M')
        commit_msg = f"Auto-update bulk pricing data - {timestamp}"
        result = git('commit', '-m', commit_msg)
        
        if result.returncode != 0:
            print(f"   [ERROR] Commit failed: {result.stderr.strip()}")
            return False
        
        # Push
        result = git('push', 'origin', 'main')
        
        if result.returncode != 0:
            print(f"   [ERROR] Push failed: {result.stderr.strip()}")
            print("   [TIP] Run this once manually to set up credentials:")
            print("      cd C:\\BulkPricing")
            print("      git push origin main")
            return False
        
        print("   [OK] Pushed to GitHub successfully!")
        return True
        
    except FileNotFoundError:
        print("   [ERROR] Git not found. Install Git for Windows: https://git-scm.com")
        return False
    except subprocess.TimeoutExpired:
        print("   [ERROR] Git command timed out (60s)")
        return False
    except Exception as e:
        print(f"   [ERROR] Git error: {e}")
        return False


def main():
    print("\n" + "="*60)
    print("  BULK PRICING GENERATOR")
    print("="*60)
    
    # Step 1: Export fresh data from CES Touch and convert
    if not import_from_ces():
        print("\n[WARN] CES export/import failed, checking for existing data...")
    
    if not os.path.exists(INPUT_FILE):
        print(f"\n[ERROR] No product data available: {INPUT_FILE}")
        print("\nTo fix this, either:")
        print("  1. Export products from CES Touch to C:\\Touch\\IMP-EXP\\sku_0002.xls")
        print("  2. Or manually place products.xlsx in C:\\BulkPricing\\data\\")
        return
    
    try:
        products = load_products(INPUT_FILE)
        
        if not products:
            print("\n[ERROR] No valid products found!")
            return
        
        print(f"\n[CALC] Calculating pricing...")
        
        products_json = []
        for p in products:
            pricing = calculate_bulk_pricing(p['cost'], p['caseqty'], p['nprice1'])
            products_json.append({
                'barcode': p['plu'],
                'name': p['desc'],
                'supplier': p['supp'],
                'supplierCode': p['suppcode'],
                'caseQty': p['caseqty'],
                'unitPrice': round(p['nprice1'], 2),
                **pricing
            })
        
        data = {
            'version': '1.0',
            'lastUpdated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'productCount': len(products_json),
            'products': products_json
        }
        
        os.makedirs(os.path.dirname(OUTPUT_FILE) or '.', exist_ok=True)
        with open(OUTPUT_FILE, 'w') as f:
            json.dump(data, f, indent=2)
        
        print(f"\n[SAVE] Saved: {OUTPUT_FILE}")
        print(f"   {len(products_json)} products")
        print(f"   {os.path.getsize(OUTPUT_FILE)/1024:.1f} KB")
        
        if GOOGLE_DRIVE_PATH:
            with open(GOOGLE_DRIVE_PATH, 'w') as f:
                json.dump(data, f, indent=2)
            print(f"   Copied to Google Drive [OK]")
        
        # Auto-push to GitHub
        if GIT_ENABLED:
            git_push(REPO_DIR, OUTPUT_FILE)
        
        print("\n[DONE] SUCCESS!\n")
        
    except Exception as e:
        print(f"\n[ERROR] ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    # Fix Windows encoding for log file output
    if sys.stdout.encoding != 'utf-8':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    main()
