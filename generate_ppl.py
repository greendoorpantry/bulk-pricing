"""
Generate ppl_0001.xls with bulk prices for CES Touch import.

Reads the existing ppl_0001.xls from C:\Touch\IMP-EXP\ and updates:
  - price2 = calculated bulk price
  - qty2 = case quantity
  - qtydesc2 = "BULK"

Leaves price0 (unused) and price1 (RRP/EACH) untouched.

Uses bulk_pricing_data.json for the calculated bulk prices.
If JSON not found, recalculates from products.xlsx.

Run from C:\BulkPricing:
  python generate_ppl.py
"""

import os
import sys
import json
import math

REPO_DIR = r"C:\BulkPricing"
JSON_FILE = os.path.join(REPO_DIR, "data", "bulk_pricing_data.json")
PRODUCTS_FILE = os.path.join(REPO_DIR, "data", "products.xlsx")
CES_PPL_SOURCE = r"C:\Touch\IMP-EXP\ppl_0001.xls"
PPL_OUTPUT = os.path.join(REPO_DIR, "data", "ppl_0001.xls")


def round_to_10c(price):
    return math.ceil(price * 10) / 10


def load_bulk_prices():
    """Load bulk prices from JSON or calculate from products.xlsx"""
    
    # Try JSON first
    if os.path.exists(JSON_FILE):
        print(f"[INFO] Loading bulk prices from {JSON_FILE}")
        with open(JSON_FILE, 'r') as f:
            data = json.load(f)
        products = data.get('products', [])
        lookup = {}
        for p in products:
            bc = str(p['barcode']).strip()
            lookup[bc] = {
                'bulk_price': p['bulkPrice'],
                'case_qty': p['caseQty'],
            }
            # Also store with leading zero stripped
            stripped = bc.lstrip('0')
            if stripped and stripped != bc:
                lookup[stripped] = lookup[bc]
        print(f"  [OK] {len(products)} products with bulk prices")
        return lookup
    
    # Fall back to calculating from products.xlsx
    if os.path.exists(PRODUCTS_FILE):
        print(f"[INFO] Calculating bulk prices from {PRODUCTS_FILE}")
        from openpyxl import load_workbook
        wb = load_workbook(PRODUCTS_FILE, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(cell.value or '').strip().lower() for cell in ws[1]]
        col_map = {h: i for i, h in enumerate(headers) if h}
        
        lookup = {}
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row: continue
            plu = str(row[col_map.get('plu', -1)] or '').strip()
            if not plu: continue
            
            cost = float(row[col_map.get('cost', -1)] or 0)
            caseqty = int(float(row[col_map.get('caseqty', -1)] or 0))
            nprice1 = float(row[col_map.get('nprice1', -1)] or 0)
            
            if cost > 0 and caseqty >= 1 and nprice1 > 0:
                bulk_price = round_to_10c(cost * caseqty * 1.1)
                lookup[plu] = {'bulk_price': bulk_price, 'case_qty': caseqty}
                stripped = plu.lstrip('0')
                if stripped and stripped != plu:
                    lookup[stripped] = lookup[plu]
                count += 1
        
        wb.close()
        print(f"  [OK] {count} products with bulk prices")
        return lookup
    
    print(f"[ERROR] Neither {JSON_FILE} nor {PRODUCTS_FILE} found")
    return {}


def main():
    print("\n" + "=" * 60)
    print("  GENERATE PPL_0001.XLS WITH BULK PRICES")
    print("  Price Level 1 = RRP/EACH (unchanged)")
    print("  Price Level 2 = Bulk price/BULK (updated)")
    print("=" * 60)
    
    # Step 1: Load bulk prices
    bulk_lookup = load_bulk_prices()
    if not bulk_lookup:
        return False
    
    # Step 2: Open existing PPL file via Excel COM
    if not os.path.exists(CES_PPL_SOURCE):
        print(f"[ERROR] PPL file not found: {CES_PPL_SOURCE}")
        return False
    
    print(f"\n[INFO] Opening {CES_PPL_SOURCE}")
    
    try:
        import win32com.client
        
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        wb = excel.Workbooks.Open(os.path.abspath(CES_PPL_SOURCE))
        ws = wb.Sheets(1)
        
        # Read headers from row 1
        headers = {}
        col = 1
        while True:
            val = ws.Cells(1, col).Value
            if val is None or str(val).strip() == '':
                break
            headers[str(val).strip().lower()] = col
            col += 1
        
        print(f"  [OK] Columns: {list(headers.keys())}")
        
        # Verify expected columns exist
        price2_col = headers.get('price2')
        qty2_col = headers.get('qty2')
        qtydesc2_col = headers.get('qtydesc2')
        cplu_col = headers.get('cplu', 1)
        
        if not all([price2_col, qty2_col, qtydesc2_col]):
            print(f"[ERROR] Missing columns: price2={price2_col}, qty2={qty2_col}, qtydesc2={qtydesc2_col}")
            wb.Close(False)
            excel.Quit()
            return False
        
        print(f"  price2=col {price2_col}, qty2=col {qty2_col}, qtydesc2=col {qtydesc2_col}")
        
        # Find last row
        last_row = ws.Cells(ws.Rows.Count, cplu_col).End(-4162).Row  # xlUp
        print(f"  [OK] {last_row - 1} product rows")
        
        # Step 3: Update price2/qty2/qtydesc2 for each product
        updated = 0
        already_ok = 0
        no_match = 0
        
        for row in range(2, last_row + 1):
            cplu = ws.Cells(row, cplu_col).Value
            if cplu is None:
                continue
            
            # Normalize barcode
            if isinstance(cplu, float):
                cplu = str(int(cplu))
            else:
                cplu = str(cplu).strip()
            
            # Look up bulk price
            bp = bulk_lookup.get(cplu) or bulk_lookup.get(cplu.lstrip('0'))
            
            if bp:
                bulk_price = bp['bulk_price']
                case_qty = bp['case_qty']
                
                # Check if already correct
                current_p2 = ws.Cells(row, price2_col).Value or 0
                current_q2 = ws.Cells(row, qty2_col).Value or 0
                current_d2 = ws.Cells(row, qtydesc2_col).Value or ''
                
                if (isinstance(current_p2, (int, float)) and 
                    abs(current_p2 - bulk_price) < 0.01 and
                    current_q2 == case_qty and
                    str(current_d2).strip() == 'BULK'):
                    already_ok += 1
                else:
                    ws.Cells(row, price2_col).Value = bulk_price
                    ws.Cells(row, qty2_col).Value = case_qty
                    ws.Cells(row, qtydesc2_col).Value = 'BULK'
                    updated += 1
            else:
                no_match += 1
        
        print(f"\n[RESULT]")
        print(f"  Updated:         {updated}")
        print(f"  Already correct: {already_ok}")
        print(f"  No bulk data:    {no_match}")
        
        # Step 4: Save as .xls (BIFF8 format)
        os.makedirs(os.path.dirname(PPL_OUTPUT) or '.', exist_ok=True)
        wb.SaveAs(os.path.abspath(PPL_OUTPUT), FileFormat=56)
        print(f"\n[SAVED] {PPL_OUTPUT}")
        
        # Copy to CES Touch import directory
        import shutil
        try:
            shutil.copy2(PPL_OUTPUT, CES_PPL_SOURCE)
            print(f"[COPIED] {CES_PPL_SOURCE}")
        except Exception as e:
            print(f"[WARN] Could not copy to CES Touch dir: {e}")
        
        wb.Close(False)
        excel.Quit()
        
        print(f"\n[DONE] PPL file ready for import!")
        print(f"  In CES Touch: Import/Export -> Import tab -> Price Levels")
        return True
        
    except ImportError:
        print("[ERROR] pywin32 not installed. Run: pip install pywin32")
        return False
    except Exception as e:
        print(f"[ERROR] {e}")
        import traceback
        traceback.print_exc()
        try:
            wb.Close(False)
            excel.Quit()
        except:
            pass
        return False


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
