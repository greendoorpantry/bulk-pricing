"""
Match IIHF catalogue with CES Touch products and update costs/case quantities.

Matching strategy:
  1. Supplier code match (CES suppcode = IIHF code) - most reliable for IIHF products
  2. Barcode match with normalization (strip leading zeros, EAN-13/UPC-A)

Usage:
  python match_iihf.py <iihf_pricelist.xlsx>
  
  e.g.: python match_iihf.py "20260210_IIHFPricelistXLR_JanFeb26.xlsx"
"""

import os
import sys
import re
from openpyxl import load_workbook


def parse_pack_qty(pack_str):
    """Extract case quantity from pack string like '6x200g', '24x130g', '50x70g'"""
    if not pack_str:
        return 0
    pack_str = str(pack_str).strip()
    match = re.match(r'(\d+)\s*[xX]', pack_str)
    if match:
        return int(match.group(1))
    return 0


def normalize_barcode(barcode):
    """Normalize barcode - strip whitespace, convert float, strip leading zeros"""
    if barcode is None:
        return ''
    if isinstance(barcode, float):
        barcode = str(int(barcode))
    else:
        barcode = str(barcode).strip()
    return barcode


def barcode_variants(barcode):
    """Generate barcode variants for matching (with/without leading zeros)"""
    bc = normalize_barcode(barcode)
    if not bc:
        return set()
    variants = {bc, bc.lstrip('0')}
    if len(bc) == 12:
        variants.add('0' + bc)  # UPC-A -> EAN-13
    if len(bc) == 13 and bc.startswith('0'):
        variants.add(bc[1:])  # EAN-13 -> UPC-A
    return variants


def load_iihf(filepath):
    """Load IIHF price list"""
    print(f"\n[IIHF] Loading: {filepath}")
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    
    # Find header row (contains 'Code', 'Description')
    header_row = None
    for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        vals = [str(v or '').strip().lower() for v in row]
        if 'code' in vals and 'description' in vals:
            header_row = row_num
            headers = [str(v or '').strip() for v in row]
            break
    
    if not header_row:
        print("  [ERROR] Could not find IIHF headers")
        return {}, {}
    
    col_map = {}
    for i, h in enumerate(headers):
        hl = h.lower()
        if hl == 'code': col_map['code'] = i
        elif hl == 'description': col_map['desc'] = i
        elif hl == 'pack': col_map['pack'] = i
        elif hl == 'price': col_map['price'] = i
        elif hl == 'retail': col_map['retail'] = i
        elif hl == 'barcode': col_map['barcode'] = i
        elif hl == 'vat': col_map['vat'] = i
    
    print(f"  [OK] Headers at row {header_row}: {list(col_map.keys())}")
    
    by_code = {}    # lookup by IIHF supplier code
    by_barcode = {} # lookup by barcode (normalized)
    
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not row[col_map.get('code', 0)]:
            continue
        
        code = str(row[col_map['code']] or '').strip()
        if not code:
            continue
        
        pack = str(row[col_map.get('pack', -1)] or '') if 'pack' in col_map else ''
        case_qty = parse_pack_qty(pack)
        case_price = float(row[col_map.get('price', -1)] or 0) if 'price' in col_map else 0
        unit_cost = case_price / case_qty if case_qty > 0 else 0
        desc = str(row[col_map.get('desc', -1)] or '') if 'desc' in col_map else ''
        barcode_raw = row[col_map.get('barcode', -1)] if 'barcode' in col_map else None
        barcode = normalize_barcode(barcode_raw)
        
        product = {
            'code': code,
            'desc': desc,
            'pack': pack,
            'case_qty': case_qty,
            'case_price': round(case_price, 4),
            'unit_cost': round(unit_cost, 4),
            'barcode': barcode,
        }
        
        # Index by supplier code (primary)
        by_code[code.upper()] = product
        
        # Index by barcode variants (secondary)
        for variant in barcode_variants(barcode):
            by_barcode[variant] = product
    
    print(f"  [OK] Loaded {len(by_code)} IIHF products by code, {len(by_barcode)} barcode entries")
    wb.close()
    return by_code, by_barcode


def load_ces_products(filepath):
    """Load CES Touch product export"""
    print(f"\n[CES] Loading: {filepath}")
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    
    headers = [str(cell.value or '').strip().lower() for cell in ws[1]]
    col_map = {h: i for i, h in enumerate(headers) if h}
    
    print(f"  [OK] Headers: {list(col_map.keys())[:15]}...")
    
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        
        plu_raw = row[col_map.get('plu', -1)]
        plu = normalize_barcode(plu_raw)
        if not plu:
            continue
        
        products.append({
            'plu': plu,
            'plu_raw': str(plu_raw or ''),
            'desc': str(row[col_map.get('desc', -1)] or ''),
            'supp': str(row[col_map.get('supp', -1)] or '').strip(),
            'suppcode': str(row[col_map.get('suppcode', -1)] or '').strip(),
            'cost': float(row[col_map.get('cost', -1)] or 0),
            'caseqty': int(row[col_map.get('caseqty', -1)] or 0),
            'nprice1': float(row[col_map.get('nprice1', -1)] or 0),
        })
    
    print(f"  [OK] Loaded {len(products)} CES products")
    iihf_count = sum(1 for p in products if p['supp'].upper() == 'IIHF')
    print(f"  [OK] {iihf_count} products with supplier = IIHF")
    wb.close()
    return products


def match_products(ces_products, iihf_by_code, iihf_by_barcode):
    """Match CES products to IIHF catalogue"""
    matched = []
    unmatched = []
    
    for prod in ces_products:
        iihf = None
        match_type = None
        
        # Strategy 1: Match by supplier code (best for IIHF products)
        if prod['supp'].upper() == 'IIHF' and prod['suppcode']:
            code = prod['suppcode'].upper()
            if code in iihf_by_code:
                iihf = iihf_by_code[code]
                match_type = 'suppcode'
        
        # Strategy 2: Match by barcode variants
        if not iihf:
            for variant in barcode_variants(prod['plu']):
                if variant in iihf_by_barcode:
                    iihf = iihf_by_barcode[variant]
                    match_type = 'barcode'
                    break
        
        if iihf:
            matched.append({**prod, 'iihf': iihf, 'match_type': match_type})
        else:
            unmatched.append(prod)
    
    return matched, unmatched


def main():
    if len(sys.argv) < 2:
        print("Usage: python match_iihf.py <iihf_pricelist.xlsx>")
        return
    
    iihf_file = sys.argv[1]
    if not os.path.exists(iihf_file):
        print(f"[ERROR] File not found: {iihf_file}")
        return
    
    # Find CES product file
    ces_file = None
    for path in ['data/products.xlsx', 'data/sku_0002.xlsx']:
        if os.path.exists(path):
            ces_file = path
            break
    
    if not ces_file:
        print("[ERROR] No CES product file found. Run generate_bulk_pricing_simple.py first.")
        return
    
    # Load data
    iihf_by_code, iihf_by_barcode = load_iihf(iihf_file)
    ces = load_ces_products(ces_file)
    
    # Match
    print("\n[MATCH] Matching products...")
    matched, unmatched = match_products(ces, iihf_by_code, iihf_by_barcode)
    
    by_suppcode = sum(1 for m in matched if m['match_type'] == 'suppcode')
    by_barcode = sum(1 for m in matched if m['match_type'] == 'barcode')
    
    print(f"\n  Matched: {len(matched)} ({by_suppcode} by supplier code, {by_barcode} by barcode)")
    print(f"  Unmatched: {len(unmatched)}")
    
    # Find products that need updates
    needs_update = []
    already_ok = []
    
    for m in matched:
        iihf_data = m['iihf']
        current_cost = m['cost']
        current_caseqty = m['caseqty']
        iihf_cost = iihf_data['unit_cost']
        iihf_caseqty = iihf_data['case_qty']
        
        cost_diff = abs(current_cost - iihf_cost) > 0.01 if current_cost > 0 else True
        qty_diff = current_caseqty != iihf_caseqty
        
        if current_cost == 0 or current_caseqty == 0 or cost_diff or qty_diff:
            needs_update.append(m)
        else:
            already_ok.append(m)
    
    print(f"\n  Need update: {len(needs_update)}")
    print(f"  Already correct: {len(already_ok)}")
    
    # Show examples
    if needs_update:
        no_cost = sum(1 for m in needs_update if m['cost'] == 0)
        no_qty = sum(1 for m in needs_update if m['caseqty'] == 0)
        cost_changed = sum(1 for m in needs_update if m['cost'] > 0 and abs(m['cost'] - m['iihf']['unit_cost']) > 0.01)
        
        print(f"\n  Breakdown:")
        print(f"    Missing cost (was 0): {no_cost}")
        print(f"    Missing caseqty (was 0): {no_qty}")
        print(f"    Cost changed: {cost_changed}")
        
        print(f"\n  First 20 updates:")
        for m in needs_update[:20]:
            i = m['iihf']
            print(f"    {m['plu']:15s} {m['desc'][:30]:30s} "
                  f"cost: {m['cost']:6.2f}->{i['unit_cost']:6.2f} "
                  f"qty: {m['caseqty']:3d}->{i['case_qty']:3d} "
                  f"[{m['match_type']}]")
    
    # Apply updates to products.xlsx
    print(f"\n[APPLY] Updating {ces_file}...")
    
    wb = load_workbook(ces_file)
    ws = wb.active
    headers = [str(cell.value or '').strip().lower() for cell in ws[1]]
    col_map = {h: i+1 for i, h in enumerate(headers) if h}
    
    # Build update lookup by PLU
    update_lookup = {}
    for m in needs_update:
        update_lookup[normalize_barcode(m['plu_raw'])] = m['iihf']
        # Also index by stripped barcode
        stripped = m['plu'].lstrip('0')
        if stripped:
            update_lookup[stripped] = m['iihf']
    
    updated = 0
    for row_num in range(2, ws.max_row + 1):
        plu_raw = ws.cell(row=row_num, column=col_map.get('plu', 1)).value
        plu = normalize_barcode(plu_raw)
        
        iihf_data = update_lookup.get(plu) or update_lookup.get(plu.lstrip('0'))
        
        if iihf_data:
            if 'cost' in col_map and iihf_data['unit_cost'] > 0:
                ws.cell(row=row_num, column=col_map['cost']).value = round(iihf_data['unit_cost'], 4)
            if 'caseqty' in col_map and iihf_data['case_qty'] > 0:
                ws.cell(row=row_num, column=col_map['caseqty']).value = iihf_data['case_qty']
            updated += 1
    
    wb.save(ces_file)
    print(f"  [OK] Updated {updated} products in {ces_file}")
    
    print(f"\n[NEXT] Run these commands to regenerate bulk pricing:")
    print(f"  python generate_bulk_pricing_simple.py")
    print(f"  python generate_ppl.py")
    print(f"\n[DONE]")


if __name__ == "__main__":
    main()
